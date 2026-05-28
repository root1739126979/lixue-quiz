from io import BytesIO
from pathlib import Path
import csv
import tempfile

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import Response
from openpyxl import load_workbook
from pydantic import BaseModel
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session, selectinload

from app.database import get_db
from app.models import (
    Admin,
    AnswerRecord,
    ExamAttempt,
    ImportBatch,
    LlmConfig,
    PointTransaction,
    PointRule,
    Question,
    QuestionBank,
    QuestionOption,
    QuestionType,
    User,
    WrongQuestion,
)
from app.security import hash_password, verify_password
from app.config import settings
from app.security import create_access_token
from app.services.dashboard import build_dashboard
from app.services.exports import rows_to_csv
from app.services.scoring import get_or_create_point_rule
from app.services.csv_importer import parse_question_csv


router = APIRouter(prefix="/api/admin", tags=["admin"])


class CreateBankRequest(BaseModel):
    name: str
    description: str | None = None
    is_active: bool = True
    exam_question_count: int = 20
    exam_time_limit_minutes: int = 30


class UpdateBankRequest(BaseModel):
    name: str | None = None
    description: str | None = None
    is_active: bool | None = None
    exam_question_count: int | None = None
    exam_time_limit_minutes: int | None = None


class UpdateQuestionRequest(BaseModel):
    stem: str | None = None
    correct_answer: str | None = None
    correct_answer_text: str | None = None
    explanation: str | None = None
    is_active: bool | None = None


class PointRulePayload(BaseModel):
    answer_base_points: int = 1
    correct_bonus_points: int = 1
    wrong_retry_correct_points: int = 1
    exam_complete_points: int = 10
    daily_point_limit: int = 100


class LlmConfigPayload(BaseModel):
    enabled: bool = False
    base_url: str | None = None
    api_key: str | None = None
    model: str | None = None
    prompt_version: str = "v1"


class AdminLoginRequest(BaseModel):
    username: str
    password: str


@router.post("/auth/login")
def admin_login(payload: AdminLoginRequest, db: Session = Depends(get_db)) -> dict:
    admin = db.scalar(select(Admin).where(Admin.username == payload.username))
    if not admin and payload.username == settings.admin_username:
        admin = Admin(username=settings.admin_username, password_hash=hash_password(settings.admin_password))
        db.add(admin)
        db.commit()
        db.refresh(admin)
    if not admin or not verify_password(payload.password, admin.password_hash):
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="账号或密码错误")
    return {
        "token": create_access_token(subject=admin.username, role="admin"),
        "admin": {"username": admin.username},
    }


@router.post("/banks", status_code=status.HTTP_201_CREATED)
def create_bank(payload: CreateBankRequest, db: Session = Depends(get_db)) -> dict:
    bank = QuestionBank(**payload.model_dump())
    db.add(bank)
    db.commit()
    db.refresh(bank)
    return bank_payload(bank)


@router.get("/banks")
def list_admin_banks(db: Session = Depends(get_db)) -> dict:
    banks = db.scalars(select(QuestionBank).order_by(QuestionBank.id)).all()
    return {"items": [bank_payload(bank) for bank in banks]}


@router.patch("/banks/{bank_id}")
def update_bank(bank_id: int, payload: UpdateBankRequest, db: Session = Depends(get_db)) -> dict:
    bank = db.get(QuestionBank, bank_id)
    if not bank:
        raise HTTPException(status_code=404, detail="题库不存在")
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(bank, key, value)
    db.commit()
    db.refresh(bank)
    return bank_payload(bank)


@router.post("/banks/{bank_id}/questions/import")
async def import_questions(
    bank_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
) -> dict:
    bank = db.get(QuestionBank, bank_id)
    if not bank:
        raise HTTPException(status_code=404, detail="题库不存在")

    content = await file.read()
    with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as temp_file:
        temp_file.write(content)
        temp_path = Path(temp_file.name)

    preview = parse_question_csv(temp_path)
    existing_source_numbers = set(
        db.scalars(select(Question.source_no).where(Question.bank_id == bank_id)).all()
    )
    duplicate_numbers = [
        parsed.source_no for parsed in preview.rows if parsed.source_no in existing_source_numbers
    ]
    if duplicate_numbers:
        raise HTTPException(status_code=400, detail=f"题号重复：{','.join(duplicate_numbers[:5])}")

    for parsed in preview.rows:
        question = Question(
            bank_id=bank_id,
            source_no=parsed.source_no,
            question_type=QuestionType(parsed.question_type),
            stem=parsed.stem,
            correct_answer=parsed.correct_answer,
            correct_answer_text=parsed.correct_answer_text,
            explanation=parsed.explanation,
        )
        question.options = [
            QuestionOption(label=option.label, content=option.content) for option in parsed.options
        ]
        db.add(question)
    db.add(
        ImportBatch(
            import_type="questions",
            total_rows=preview.total_rows,
            imported_count=preview.valid_count,
            error_count=len(preview.errors),
        )
    )
    db.commit()
    return {
        "total_rows": preview.total_rows,
        "imported_count": preview.valid_count,
        "type_counts": preview.type_counts,
        "errors": [error.__dict__ for error in preview.errors],
    }


@router.get("/questions")
def list_questions(bank_id: int | None = None, db: Session = Depends(get_db)) -> dict:
    query = select(Question).options(selectinload(Question.options)).order_by(Question.id)
    if bank_id is not None:
        query = query.where(Question.bank_id == bank_id)
    questions = db.scalars(query).all()
    return {"items": [question_payload(question, include_answer=True) for question in questions]}


@router.get("/questions/{question_id}")
def get_question(question_id: int, db: Session = Depends(get_db)) -> dict:
    question = db.scalar(
        select(Question).options(selectinload(Question.options)).where(Question.id == question_id)
    )
    if not question:
        raise HTTPException(status_code=404, detail="题目不存在")
    return question_payload(question, include_answer=True)


@router.patch("/questions/{question_id}")
def update_question(
    question_id: int,
    payload: UpdateQuestionRequest,
    db: Session = Depends(get_db),
) -> dict:
    question = db.get(Question, question_id)
    if not question:
        raise HTTPException(status_code=404, detail="题目不存在")
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(question, key, value)
    db.commit()
    db.refresh(question)
    return question_payload(question, include_answer=True)


@router.post("/employees/import")
async def import_employees(file: UploadFile = File(...), db: Session = Depends(get_db)) -> dict:
    content = await file.read()
    rows = _read_employee_rows(content, file.filename or "")
    seen_work_numbers: set[str] = set()
    seen_phones: set[str] = set()
    errors: list[dict] = []
    parsed_rows: list[dict] = []

    for row_number, row in enumerate(rows, start=2):
        name = (row.get("姓名") or "").strip()
        work_no = (row.get("工号") or "").strip() or None
        phone = (row.get("手机号") or "").strip() or None
        department = (row.get("部门") or "").strip() or None
        status_text = (row.get("状态") or "启用").strip()
        password = (row.get("初始密码") or "").strip()
        if not name or not (work_no or phone):
            errors.append({"row_number": row_number, "message": "姓名以及工号或手机号不能为空"})
            continue
        if work_no and work_no in seen_work_numbers:
            raise HTTPException(status_code=400, detail=f"工号重复：{work_no}")
        if phone and phone in seen_phones:
            raise HTTPException(status_code=400, detail=f"手机号重复：{phone}")
        if work_no:
            seen_work_numbers.add(work_no)
        if phone:
            seen_phones.add(phone)
        parsed_rows.append(
            {
                "name": name,
                "work_no": work_no,
                "phone": phone,
                "department": department,
                "is_active": status_text != "停用",
                "password": password or "123456",
                "password_provided": bool(password),
            }
        )

    imported_users: list[User] = []
    for row in parsed_rows:
        conditions = []
        if row["work_no"]:
            conditions.append(User.work_no == row["work_no"])
        if row["phone"]:
            conditions.append(User.phone == row["phone"])
        existing = db.scalar(select(User).where(or_(*conditions))) if conditions else None
        if existing:
            existing.name = row["name"]
            existing.work_no = row["work_no"]
            existing.phone = row["phone"]
            existing.department = row["department"]
            existing.is_active = row["is_active"]
            if row["password_provided"]:
                existing.password_hash = hash_password(row["password"])
            imported_users.append(existing)
        else:
            user = User(
                name=row["name"],
                work_no=row["work_no"],
                phone=row["phone"],
                department=row["department"],
                is_active=row["is_active"],
                password_hash=hash_password(row["password"]),
            )
            db.add(user)
            imported_users.append(user)
    db.add(
        ImportBatch(
            import_type="employees",
            total_rows=len(rows),
            imported_count=len(parsed_rows),
            error_count=len(errors),
        )
    )
    db.commit()
    for user in imported_users:
        db.refresh(user)
    return {
        "imported_count": len(parsed_rows),
        "errors": errors,
        "users": [
            {
                "id": user.id,
                "name": user.name,
                "work_no": user.work_no,
                "phone": user.phone,
                "department": user.department,
                "is_active": user.is_active,
            }
            for user in imported_users
        ],
    }


@router.get("/employees")
def list_employees(db: Session = Depends(get_db)) -> dict:
    users = db.scalars(select(User).order_by(User.id)).all()
    return {
        "items": [
            {
                "id": user.id,
                "name": user.name,
                "work_no": user.work_no,
                "phone": user.phone,
                "department": user.department,
                "is_active": user.is_active,
            }
            for user in users
        ]
    }


@router.get("/employees/{employee_id}/summary")
def get_employee_summary(employee_id: int, db: Session = Depends(get_db)) -> dict:
    user = db.get(User, employee_id)
    if not user:
        raise HTTPException(status_code=404, detail="员工不存在")

    answer_count = (
        db.scalar(select(func.count(AnswerRecord.id)).where(AnswerRecord.user_id == user.id)) or 0
    )
    correct_count = (
        db.scalar(
            select(func.count(AnswerRecord.id)).where(
                AnswerRecord.user_id == user.id,
                AnswerRecord.is_correct.is_(True),
            )
        )
        or 0
    )
    total_points = (
        db.scalar(
            select(func.coalesce(func.sum(PointTransaction.points), 0)).where(
                PointTransaction.user_id == user.id
            )
        )
        or 0
    )
    open_wrong_count = (
        db.scalar(
            select(func.count(WrongQuestion.id)).where(
                WrongQuestion.user_id == user.id,
                WrongQuestion.status == "open",
            )
        )
        or 0
    )
    return {
        "id": user.id,
        "name": user.name,
        "work_no": user.work_no,
        "phone": user.phone,
        "department": user.department,
        "is_active": user.is_active,
        "answer_count": answer_count,
        "accuracy": round(correct_count * 100 / answer_count, 2) if answer_count else 0,
        "total_points": int(total_points),
        "open_wrong_count": open_wrong_count,
    }


@router.get("/point-rules")
def get_point_rules(db: Session = Depends(get_db)) -> dict:
    return point_rule_payload(get_or_create_point_rule(db))


@router.put("/point-rules")
def update_point_rules(payload: PointRulePayload, db: Session = Depends(get_db)) -> dict:
    rule = get_or_create_point_rule(db)
    for key, value in payload.model_dump().items():
        setattr(rule, key, value)
    db.commit()
    db.refresh(rule)
    return point_rule_payload(rule)


@router.get("/dashboard")
def dashboard(db: Session = Depends(get_db)) -> dict:
    return build_dashboard(db)


@router.get("/llm-config")
def get_llm_config(db: Session = Depends(get_db)) -> dict:
    config = db.scalar(select(LlmConfig).limit(1))
    if not config:
        return {
            "enabled": False,
            "base_url": "",
            "api_key_masked": "",
            "model": "",
            "prompt_version": "v1",
        }
    return {
        "enabled": config.enabled,
        "base_url": config.base_url or "",
        "api_key_masked": "********" if config.api_key_encrypted else "",
        "model": config.model or "",
        "prompt_version": config.prompt_version,
    }


@router.put("/llm-config")
def update_llm_config(payload: LlmConfigPayload, db: Session = Depends(get_db)) -> dict:
    config = db.scalar(select(LlmConfig).limit(1))
    if not config:
        config = LlmConfig()
        db.add(config)
    config.enabled = payload.enabled
    config.base_url = payload.base_url
    config.model = payload.model
    config.prompt_version = payload.prompt_version
    if payload.api_key:
        config.api_key_encrypted = payload.api_key
    db.commit()
    return get_llm_config(db)


@router.get("/exports/users")
def export_users(db: Session = Depends(get_db)) -> Response:
    users = db.scalars(select(User).order_by(User.id)).all()
    rows = []
    for user in users:
        answer_count = len(
            db.scalars(select(AnswerRecord).where(AnswerRecord.user_id == user.id)).all()
        )
        rows.append([user.name, user.work_no or "", user.phone or "", user.department or "", answer_count])
    return csv_response(rows_to_csv(["姓名", "工号", "手机号", "部门", "累计答题数"], rows))


@router.get("/exports/answers")
def export_answers(db: Session = Depends(get_db)) -> Response:
    answers = db.scalars(select(AnswerRecord).order_by(AnswerRecord.id)).all()
    rows = [
        [
            answer.user_id,
            answer.bank_id,
            answer.question_id,
            answer.selected_answer,
            "是" if answer.is_correct else "否",
            answer.points_awarded,
        ]
        for answer in answers
    ]
    return csv_response(rows_to_csv(["用户ID", "题库ID", "题目ID", "选择答案", "是否正确", "积分"], rows))


@router.get("/exports/exams")
def export_exams(db: Session = Depends(get_db)) -> Response:
    attempts = db.scalars(select(ExamAttempt).order_by(ExamAttempt.id)).all()
    rows = [
        [
            attempt.user_id,
            attempt.bank_id,
            attempt.question_count,
            attempt.score or "",
            attempt.correct_count or "",
            attempt.submitted_at or "",
        ]
        for attempt in attempts
    ]
    return csv_response(rows_to_csv(["用户ID", "题库ID", "题数", "分数", "正确数", "提交时间"], rows))


def bank_payload(bank: QuestionBank) -> dict:
    return {
        "id": bank.id,
        "name": bank.name,
        "description": bank.description,
        "is_active": bank.is_active,
        "exam_question_count": bank.exam_question_count,
        "exam_time_limit_minutes": bank.exam_time_limit_minutes,
    }


def question_payload(question: Question, *, include_answer: bool) -> dict:
    payload = {
        "id": question.id,
        "bank_id": question.bank_id,
        "source_no": question.source_no,
        "question_type": question.question_type.value,
        "stem": question.stem,
        "options": [
            {"label": option.label, "content": option.content}
            for option in sorted(question.options, key=lambda item: item.label)
        ],
        "is_active": question.is_active,
    }
    if include_answer:
        payload.update(
            {
                "correct_answer": question.correct_answer,
                "correct_answer_text": question.correct_answer_text,
                "explanation": question.explanation,
            }
        )
    return payload


def point_rule_payload(rule: PointRule) -> dict:
    return {
        "answer_base_points": rule.answer_base_points,
        "correct_bonus_points": rule.correct_bonus_points,
        "wrong_retry_correct_points": rule.wrong_retry_correct_points,
        "exam_complete_points": rule.exam_complete_points,
        "daily_point_limit": rule.daily_point_limit,
    }


def _read_employee_rows(content: bytes, filename: str) -> list[dict]:
    if filename.lower().endswith(".xlsx") or content.startswith(b"PK\x03\x04"):
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [_cell_text(value) for value in rows[0]]
        return [
            {headers[index]: _cell_text(value) for index, value in enumerate(row) if index < len(headers)}
            for row in rows[1:]
            if any(value is not None and str(value).strip() for value in row)
        ]
    text = content.decode("utf-8-sig")
    return [
        {key: _cell_text(value) for key, value in row.items()}
        for row in csv.DictReader(text.splitlines())
    ]


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def csv_response(content: str) -> Response:
    return Response(content=content, media_type="text/csv; charset=utf-8")
